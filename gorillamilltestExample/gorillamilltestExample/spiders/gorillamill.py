import csv
import os
import scrapy
from openpyxl import Workbook, load_workbook


class GorillaMillSpider(scrapy.Spider):
    name = "gorillamill"
    allowed_domains = ["gorillamill.com"]

    def start_requests(self):
        with open("Gorilla-Mill-Products.csv", 'r') as file:
            reader = csv.reader(file)
            next(reader)  # skip header row
            for row in reader:
                try:
                    product_id = row[0]
                    url = f"https://gorillamill.com/products/product/{product_id}"
                    yield scrapy.Request(url, self.parse, meta={"product_id": product_id})
                except:
                    product_id = 'Id not working'

    def parse(self, response):
        itemlist = {}
        title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
        title = title.strip().replace('\n', '').replace("\r", "").replace("\t", "") if title else ''
        # image url
        image_url = response.css('div#specs img.tool::attr(src)').get()
        # catalog_number 
        catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
        # price
        price = response.xpath('//strong[text()="Price(USD):"]/following-sibling::text()').get().strip()
        # check stock
        checkstock = response.css('a[data-stock]::attr(data-stock)').get()

        # attributes
        desc = response.xpath('normalize-space(//div[@class="uk-width-1-1 uk-width-medium-4-10"])')
        attributes = desc.get().replace('Check Stock >>', '').replace(',', '| ')  if desc else ''
        
        desc_element = response.xpath('//div[@id="desc"]')
        desc_html = ''.join(desc_element.extract())
        remove_elements = response.xpath('//img[@class="brand"] | //a')
        for img_element in remove_elements:
            img_text = img_element.extract()
            desc_html = desc_html.replace(img_text, '').replace('""', '')
        tolerances_html = response.xpath('//*[@id="tolerances"]')
        description = desc_html + ''.join(tolerances_html.extract())
        
        diagram_url = response.css('div#diagram-inner img.uk-vertical-align-middle::attr(src)').get()        


        itemlist = {
            "catalog_number": catalog_number,
            "title": title,
            'attributes': attributes,
            'description': description,
            "price": price,
            "Check Stock": checkstock,
            "image_url": image_url,
            'diagram_url': diagram_url
        }
        print(itemlist)

        # Load workbook if exist, or create a new one
        if os.path.exists('Gorilla Mill example output.xlsx'):
            book = load_workbook('Gorilla Mill example output.xlsx')
        else:
            book = Workbook()

        # Get sheet and append data
        sheet_name = "Gorilla Mill Products"
        if sheet_name not in book.sheetnames:
            sheet = book.active
            sheet.title = sheet_name
            # add header row
            header = ["Product ID", "catalog_number", "title", "attributes", "description", "price", "Check Stock", "image_url",  "diagram_url"]
            sheet.append(header)
        else:
            sheet = book[sheet_name]

        # add data row
        product_id = response.meta.get('product_id')
        row = [product_id, itemlist['catalog_number'], itemlist['title'], itemlist['attributes'],itemlist['description'], itemlist['price'], itemlist['Check Stock'], itemlist['image_url'],itemlist['diagram_url']]
        sheet.append(row)

        # Save the workbook
        try:
            book.save('Gorilla Mill example output.xlsx')
        except ValueError:
            for idx, val in enumerate(row):
                row[idx] = str(val)
            sheet.append(row)
            book.save('Gorilla Mill example output.xlsx')
